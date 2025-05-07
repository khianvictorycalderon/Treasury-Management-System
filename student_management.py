import tkinter as tk
from tkinter import messagebox
from openpyxl import Workbook, load_workbook
import os
from bg_images import *
from bg_loader import load_bg_image

BLUE = "#4190F3"
YELLOW = "#F5B940"
DARK_BLUE = "#17355A"
GREEN = "#66B857"
WHITE = "#FFFFFF"
BG_ENTRY = "#F7FAFC"
RED = "#F46E6E"

student_list = []

def save_all_students_to_excel():
    try:
        if not os.path.exists("userdata.xlsx"):
            wb = Workbook()
            sheet = wb.active
            sheet.title = "Student_Management"
            sheet.append(["ID", "First Name", "Middle Initial", "Last Name"])
        else:
            wb = load_workbook("userdata.xlsx")

        if "Student_Management" not in wb.sheetnames:
            sheet = wb.create_sheet("Student_Management")
            sheet.append(["ID", "First Name", "Middle Initial", "Last Name"])
        else:
            sheet = wb["Student_Management"]

        # Clear all data except for the header
        sheet.delete_rows(2, sheet.max_row)

        # Add all students from the list to the sheet
        for student in student_list:
            sheet.append([student['id'], student['first_name'], student['middle_initial'], student['last_name']])

        wb.save("userdata.xlsx")

    except Exception as e:
        messagebox.showerror("Error", f"Failed to save students: {e}")

label_opts = {"font": ("Segoe UI", 14, "bold"), "fg": DARK_BLUE, "bg": WHITE}
entry_opts = {
    "font": ("Segoe UI", 14),
    "fg": DARK_BLUE,
    "bg": BG_ENTRY,
    "highlightthickness": 2,
    "highlightbackground": DARK_BLUE,
    "relief": tk.FLAT,
    "bd": 0
}
   
def create_student_management_page(parent):
    
    global student_list

    outer_frame = tk.Frame(parent,bg=DARK_BLUE)
    outer_frame.pack(fill="both", expand=True)
    outer_frame.grid_rowconfigure(0, weight=1)
    outer_frame.grid_columnconfigure(0, weight=1)

    frame = tk.Frame(outer_frame,width=800, height=300, bg=WHITE, highlightthickness=1, highlightbackground=DARK_BLUE)
    frame.grid(row=0, column=0)
    
    canvas = tk.Canvas(outer_frame, bg=WHITE, highlightthickness=200, highlightbackground=DARK_BLUE)
    canvas.place(x=0, y=0, relwidth=1, relheight=1)  # Cover the entire frame area
    load_bg_image(canvas, STUDENT_MANAGEMENT_BACKGROUND_IMAGE)  # Load the background image

    # Ensure that canvas is placed in the correct stacking order
    canvas.tk.call("lower", canvas._w, None)

    tk.Label(frame, text="Student Management", font=("Segoe UI", 40, "bold"), fg=BLUE, bg=WHITE).grid(row=0, column=0, columnspan=2, pady=(30, 20))
    
    tk.Label(frame, text="ID:", **label_opts).grid(row=1, column=0, sticky="e", padx=10, pady=5)
    entry_id = tk.Entry(frame, width=30, **entry_opts)
    entry_id.grid(row=1, column=1, pady=5)

    tk.Label(frame, text="First Name:", **label_opts).grid(row=2, column=0, sticky="e", padx=10, pady=5)
    entry_fname = tk.Entry(frame, width=30, **entry_opts)
    entry_fname.grid(row=2, column=1, pady=5)

    tk.Label(frame, text="Middle Initial:", **label_opts).grid(row=3, column=0, sticky="e", padx=10, pady=5)
    entry_mname = tk.Entry(frame, width=30, **entry_opts)
    entry_mname.grid(row=3, column=1, pady=5)


    tk.Label(frame, text="Last Name:", **label_opts).grid(row=4, column=0, sticky="e", padx=10, pady=5)
    entry_lname = tk.Entry(frame, width=30, **entry_opts)
    entry_lname.grid(row=4, column=1, pady=5)
    def on_add_enter(e): add_btn.config(bg="#007A17")
    def on_add_leave(e): add_btn.config(bg="#009820")
    
    add_btn = tk.Button(
    frame,
    text="Add Student",
    font=("Segoe UI", 14, "bold"),
    bg="#009820",                # Green background
    fg="white",                  # White text
    highlightthickness=2,
    highlightbackground=DARK_BLUE,  # Blue border
    activebackground="#007A17",  # Darker green on press/hover
    activeforeground="white",
    bd=0,
    relief=tk.FLAT,
    cursor="hand2",
    command=lambda: add_student(entry_id, entry_fname, entry_mname, entry_lname)
)
    add_btn.grid(row=5, column=0, columnspan=2, pady=(20, 10))
    add_btn.bind("<Enter>", on_add_enter)
    add_btn.bind("<Leave>", on_add_leave)
   
    list_container = tk.Frame(frame, bg=WHITE)
    list_container.grid(row=6, column=0, columnspan=2, pady=20)

    # Canvas and Scrollbar setup
    canvas = tk.Canvas(list_container, width=550, height=300, bg=WHITE, highlightthickness=2, highlightbackground=DARK_BLUE)
    scrollbar = tk.Scrollbar(list_container, orient="vertical", command=canvas.yview, width=25)
    scrollable_frame = tk.Frame(canvas, bg=WHITE)

    # Bind the configure event to update the scroll region
    scrollable_frame.bind("<Configure>", lambda e: canvas.configure(scrollregion=canvas.bbox("all")))

    # Create window in canvas for the scrollable frame
    canvas.create_window((0, 0), window=scrollable_frame, anchor="nw")
    canvas.configure(yscrollcommand=scrollbar.set)

    # Pack canvas and scrollbar properly
    canvas.pack(side="left", fill="both", expand=True)
    scrollbar.pack(side="right", fill="y")

    # Mousewheel event binding to scrollable frame
    def on_mousewheel(event):
        canvas.yview_scroll(int(-1 * (event.delta / 120)), "units")

    # Bind the mousewheel event to the canvas and scrollable frame
    canvas.bind("<Enter>", lambda e: canvas.bind_all("<MouseWheel>", on_mousewheel))
    canvas.bind("<Leave>", lambda e: canvas.unbind_all("<MouseWheel>"))
    scrollable_frame.bind_all("<MouseWheel>", on_mousewheel)  # Ensure scroll also works when hovering the list area


    # Inside refresh_student_list function:
    def refresh_student_list():
        for widget in scrollable_frame.winfo_children():
            widget.destroy()

        sorted_students = sorted(student_list, key=lambda s: (s['last_name'].lower(), s['first_name'].lower()))

        for index, student in enumerate(sorted_students):
            row_widgets = {}

            id_entry = tk.Entry(scrollable_frame, width=10, font=("Arial", 12), bg=WHITE, fg=DARK_BLUE, bd=1, relief="solid", justify="center")
            id_entry.insert(0, student['id'])
            id_entry.config(state='disabled')
            id_entry.grid(row=index, column=0, padx=5, pady=5)
            row_widgets['id'] = id_entry

            lname_entry = tk.Entry(scrollable_frame, width=15, font=("Arial", 12), bg=WHITE, fg=DARK_BLUE, bd=1, relief="solid", justify="center")
            lname_entry.insert(0, student['last_name'])
            lname_entry.config(state='disabled')
            lname_entry.grid(row=index, column=1, padx=5, pady=5)
            row_widgets['last_name'] = lname_entry

            fname_entry = tk.Entry(scrollable_frame, width=15, font=("Arial", 12), bg=WHITE, fg=DARK_BLUE, bd=1, relief="solid", justify="center")
            fname_entry.insert(0, student['first_name'])
            fname_entry.config(state='disabled')
            fname_entry.grid(row=index, column=2, padx=5, pady=5)
            row_widgets['first_name'] = fname_entry

            mname_entry = tk.Entry(scrollable_frame, width=5, font=("Arial", 12), bg=WHITE, fg=DARK_BLUE, bd=1, relief="solid", justify="center")
            mname_entry.insert(0, student['middle_initial'])
            mname_entry.config(state='disabled')
            mname_entry.grid(row=index, column=3, padx=5, pady=5)
            row_widgets['middle_initial'] = mname_entry

            def make_edit_save_buttons(i, row, s):
                editing = {'mode': False}

                def toggle_edit_save():
                    if not editing['mode']:
                        for entry in row.values():
                            if isinstance(entry, tk.Entry):
                                entry.config(state='normal')
                        row['edit_button'].config(text="Save")
                        editing['mode'] = True
                    else:
                        new_id = row['id'].get().strip()
                        new_fname = row['first_name'].get().strip()
                        new_mname = row['middle_initial'].get().strip()
                        new_lname = row['last_name'].get().strip()

                        if not (new_id and new_fname and new_lname):
                            messagebox.showwarning("Input Error", "ID, First Name, and Last Name are required.")
                            return

                        for other in student_list:
                            if other != s and other['id'] == new_id:
                                messagebox.showerror("Duplicate ID", "Another student already has this ID.")
                                return

                        s['id'] = new_id
                        s['first_name'] = new_fname
                        s['middle_initial'] = new_mname
                        s['last_name'] = new_lname

                        save_all_students_to_excel()

                        for entry in row.values():
                            if isinstance(entry, tk.Entry):
                                entry.config(state='disabled')

                        row['edit_button'].config(text="Edit")
                        editing['mode'] = False

                return toggle_edit_save

            edit_button = tk.Button(scrollable_frame, text="Edit", font=("Arial", 10),fg=WHITE, bg=GREEN)
            edit_button.grid(row=index, column=4, padx=2, pady=2)
            row_widgets['edit_button'] = edit_button
            edit_button.config(command=make_edit_save_buttons(index, row_widgets, student))

            delete_button = tk.Button(scrollable_frame, text="Delete", font=("Arial", 10,), command=lambda i=index: delete_student(i), fg=WHITE, bg=RED)
            delete_button.grid(row=index, column=5, padx=2, pady=2)

        # Update scrollregion for the canvas to enable scrolling
        canvas.configure(scrollregion=canvas.bbox("all"))

    def add_student(entry_id, entry_fname, entry_mname, entry_lname):
        id_value = entry_id.get().strip()
        first_name = entry_fname.get().strip()
        middle_initial = entry_mname.get().strip()
        last_name = entry_lname.get().strip()

        if not (id_value and first_name and last_name):
            messagebox.showwarning("Input Error", "ID, First Name, and Last Name are required.")
            return

        # Check for duplicate ID in memory (student_list)
        for student in student_list:
            if student['id'] == id_value:
                messagebox.showerror("Duplicate ID", "A student with this ID already exists.")
                return

        # Add the new student to the in-memory list
        student_list.append({
            'id': id_value,
            'first_name': first_name,
            'middle_initial': middle_initial,
            'last_name': last_name
        })

        # Save to Excel
        save_all_students_to_excel()

        # Refresh displayed student list from memory (NOT Excel)
        refresh_student_list()

        # Clear input fields
        clear_fields()



    def delete_student(index):
        sorted_students = sorted(student_list, key=lambda s: (s['last_name'].lower(), s['first_name'].lower()))
        student_to_delete = sorted_students[index]
        student_list.remove(student_to_delete)
        save_all_students_to_excel()
        refresh_student_list()

    def clear_fields():
        entry_id.delete(0, tk.END)
        entry_fname.delete(0, tk.END)
        entry_mname.delete(0, tk.END)
        entry_lname.delete(0, tk.END)

    def load_students_from_excel():
        global student_list
        student_list.clear()
        if not os.path.exists("userdata.xlsx"):
            return

        try:
            wb = load_workbook("userdata.xlsx")
            if "Student_Management" not in wb.sheetnames:
                return

            sheet = wb["Student_Management"]
            for row in sheet.iter_rows(min_row=2, values_only=True):
                if row[0] is None:
                    continue
                student_list.append({
                    'id': str(row[0]),
                    'first_name': str(row[1]),
                    'middle_initial': str(row[2]),
                    'last_name': str(row[3])
                })
        except Exception as e:
            messagebox.showerror("Error", f"Failed to load students from Excel: {e}")

    load_students_from_excel()
    refresh_student_list()

    return outer_frame