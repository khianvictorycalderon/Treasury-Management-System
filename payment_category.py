from tkinter import *
from tkinter import messagebox
import tkinter as tk
from openpyxl import Workbook, load_workbook
import os
from bg_images import *
from bg_loader import load_bg_image

BLUE = "#4190F3"
YELLOW = "#F5B940"
DARK_BLUE = "#17355A"
GREEN = "#66B857"
WHITE = "#FFFFFF"
BG_ENTRY = "#F7FAFC"
RED = "#F46E6E" 

category_list = []
EXCEL_FILE = 'userdata.xlsx'
CATEGORY_SHEET = 'Payment_Categories'

def load_from_excel():
    global category_list
    category_list.clear()
    if os.path.exists(EXCEL_FILE):
        try:
            wb = load_workbook(EXCEL_FILE)
            if CATEGORY_SHEET in wb.sheetnames:
                sheet = wb[CATEGORY_SHEET]

                for row in sheet.iter_rows(min_row=1, values_only=True):
                    if row[0] is None:
                        continue
                    name = str(row[0]).strip()
                    try:
                        fund = float(row[1])
                    except Exception:
                        fund = 0
                    if not any(cat[0].lower() == name.lower() for cat in category_list):
                        category_list.append((name, fund))
        except Exception as e:
            messagebox.showerror("Error Loading Excel", f"Could not load Excel: {e}")

def export_to_excel():
    try:
        if os.path.exists(EXCEL_FILE):
            wb = load_workbook(EXCEL_FILE)
        else:
            wb = Workbook()

        if CATEGORY_SHEET in wb.sheetnames:
            sheet = wb[CATEGORY_SHEET]
        else:
            sheet = wb.create_sheet(CATEGORY_SHEET)
            sheet.append(['Payment Category Name', 'Required Fund (Per Student)'])

        if sheet.max_row > 1:
            sheet.delete_rows(2, sheet.max_row - 1)

        for name, fund in category_list:
            sheet.append([name, fund])

        wb.save(EXCEL_FILE)

    except Exception as e:
        messagebox.showerror("Error", f"Failed to export to Excel: {e}")

def create_payment_category_page(parent):
    global category_list
    outer_frame = Frame(parent,bg=WHITE, highlightthickness=1, highlightbackground=DARK_BLUE)
    outer_frame.pack(fill="both", expand=True)
    outer_frame.grid_rowconfigure(0, weight=1)
    outer_frame.grid_columnconfigure(0, weight=1)

    frame = Frame(outer_frame,bg=WHITE, highlightthickness=1, highlightbackground=DARK_BLUE)
    frame.grid(row=0, column=0)
    
    canvas = tk.Canvas(outer_frame)
    canvas.place(x=0, y=0, relwidth=1, relheight=1)  # Cover the entire frame area
    load_bg_image(canvas, PAYMENT_CATEGORY_BACKGROUND_IMAGE)  # Load the background image

    # Ensure that canvas is placed in the correct stacking order
    canvas.tk.call("lower", canvas._w, None)

    Label(frame, text="Payment Category Page",
    font=("Segoe UI", 40, "bold"),
    fg=BLUE, bg=WHITE
    ).grid(row=0, column=0, columnspan=4, pady=(30, 20))

    input_frame = tk.Frame(frame, bg=WHITE)
    input_frame.grid(row=1, column=0, columnspan=4, pady=10)
    
    label_opts = {"font": ("Segoe UI", 14, "bold"), "fg": DARK_BLUE, "bg": WHITE}
    entry_opts = {
    "font": ("Segoe UI", 14),
    "fg": DARK_BLUE,
    "bg": BG_ENTRY,
    "highlightthickness": 2,
    "highlightbackground": BLUE,
    "relief": tk.FLAT,
    "bd": 0
}

    Label(input_frame, text="Payment Category Name:", **label_opts).grid(row=0, column=0, padx=10, pady=5, sticky="e")
    entry_name = tk.Entry(input_frame, width=30, **entry_opts)
    entry_name.grid(row=0, column=1, padx=10, pady=5)

    Label(input_frame, text="Required Fund (Per Student):", **label_opts).grid(row=1, column=0, padx=10, pady=5, sticky="e")
    entry_fund = tk.Entry(input_frame, width=30, **entry_opts)
    entry_fund.grid(row=1, column=1, padx=10, pady=5)
    
    headers = tk.Frame(frame, bg=WHITE)
    tk.Label(headers, text="Payment Category Name", font=("Segoe UI", 14, "bold"),
         fg=WHITE, bg=BLUE, borderwidth=2, relief="solid", width=40, anchor="center", justify="center")\
    .grid(row=0, column=0, padx=2, pady=2)
    tk.Label(headers, text="Required Fund", font=("Segoe UI", 14, "bold"),
         fg=WHITE, bg=BLUE, borderwidth=2, relief="solid", width=25, anchor="center", justify="center")\
    .grid(row=0, column=1, padx=2, pady=2)

    
    headers.grid(row=2, column=0, columnspan=2) # Make way for the delete button

    # Create a frame to hold the table and scrollbar
    frame_table = Frame(frame, width=500,bg=WHITE)
    frame_table.grid(row=4, column=0, columnspan=4, pady=10, sticky="nsew")

    # Create a canvas to allow scrolling
    scroll_canvas = tk.Canvas(frame_table,bg=WHITE)
    scroll_canvas.pack(side="left", fill="both", expand=True)
    
    def on_mousewheel(event):
        scroll_canvas.yview_scroll(int(-1 * (event.delta / 120)), "units")

    scroll_canvas.bind_all("<MouseWheel>", on_mousewheel)

    # Create a vertical scrollbar linked to the canvas (reduced width for proper display)
    scrollbar = tk.Scrollbar(frame_table, orient="vertical", command=scroll_canvas.yview, width=15)
    scrollbar.pack(side="right", fill="y")

    # Link the scrollbar to the canvas
    scroll_canvas.configure(yscrollcommand=scrollbar.set)

    # Create a frame that will hold the table inside the canvas
    table_frame = Frame(scroll_canvas)

    # Add the table_frame to the canvas with a window
    scroll_canvas.create_window((0, 0), window=table_frame, anchor="nw")


    # Make the table scrollable by updating the scroll region of the canvas
    def refresh_table():
        for widget in table_frame.winfo_children():
            widget.destroy()

        # Populate the table with payment categories
        for i, (name, fund) in enumerate(category_list):
            Label(table_frame, text=name, font=("Arial", 12), width=50, anchor="w",
                borderwidth=1, relief="solid").grid(row=i + 1, column=0, padx=2, pady=2)
            Label(table_frame, text=str(fund), font=("Arial", 12), width=20, anchor="w",
                borderwidth=1, relief="solid").grid(row=i + 1, column=1, padx=2, pady=2)

            def make_delete_handler(index):
                return lambda: (
                    category_list.pop(index),
                    export_to_excel(),
                    refresh_table()
                ) if messagebox.askyesno("Confirm Delete", f"Delete category '{category_list[index][0]}'?") else None

            Button(table_frame, text="Delete",
                font=("Segoe UI", 10, "bold"),
                bg=RED, fg=WHITE,
                activebackground="#C93030", activeforeground=WHITE,
                bd=0, relief=tk.FLAT, cursor="hand2",
                command=make_delete_handler(i))\
                .grid(row=i + 1, column=2, padx=2, pady=2)

        scroll_canvas.update_idletasks()
        scroll_canvas.config(scrollregion=scroll_canvas.bbox("all"))

    
    def add_category():
        name = entry_name.get().strip()
        fund = entry_fund.get().strip()

        if not name or not fund:
            messagebox.showerror("Input Error", "Both fields are required.")
            return

        try:
            fund_value = float(fund)
            if fund_value <= 0:
                raise ValueError
        except ValueError:
            messagebox.showerror("Input Error", "Required fund must be a positive number.")
            return

        if any(cat[0].lower() == name.lower() for cat in category_list):
            messagebox.showwarning("Duplicate Entry", "This category already exists.")
            return

        category_list.append((name, fund_value))
        export_to_excel()
        refresh_table()

        entry_name.delete(0, END)
        entry_fund.delete(0, END)

    btn_add = tk.Button(
        frame,
        text="Add Category",
        font=("Segoe UI", 14, "bold"),
        bg="#009820",
        fg="white",
        highlightthickness=2,
        highlightbackground="#4190F3",
        activebackground="#007A17",
        activeforeground="white",
        bd=0,
        relief=tk.FLAT,
        cursor="hand2",
        command=add_category
    )
    btn_add.grid(row=3, column=0, columnspan=4, pady=(10, 20))

    load_from_excel()
    refresh_table()

    def auto_refresh():
        load_from_excel()
        refresh_table()
        outer_frame.after(1000, auto_refresh)

    auto_refresh()

    return outer_frame
