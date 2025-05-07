import tkinter as tk
from tkinter import ttk, messagebox
from datetime import datetime
import openpyxl
import os
from bg_images import *
from bg_loader import load_bg_image

BLUE = "#4190F3"
YELLOW = "#F5B940"
DARK_BLUE = "#17355A"
GREEN = "#009820"
WHITE = "#FFFFFF"
BG_ENTRY = "#F7FAFC"


EXCEL_FILE = "userdata.xlsx"
STUDENT_SHEET = "Student_Management"
CATEGORY_SHEET = "Payment_Categories"
RECORD_SHEET = "Payment_Records"

def create_payment_record_page(parent):
   # Modern background color (white) for a clean look
    frame = tk.Frame(parent, bg=WHITE ,highlightthickness=2, highlightbackground=DARK_BLUE)
    frame.pack(fill="both", expand=True)

    # Canvas with no border, covers the whole frame
    canvas = tk.Canvas(frame, bg=WHITE, highlightthickness=1, highlightbackground=WHITE)
    canvas.pack(fill="both", expand=True)   
    canvas.place(x=0, y=0, relwidth=1, relheight=1)  # Cover the entire frame area

    # Load your background image as before
    load_bg_image(canvas, PAYMENT_RECORD_BACKGROUND_IMAGE)

    tk.Label(frame, text="Payment Record Page",font=("Segoe UI", 32, "bold"),fg="#4190F3",  bg="#FFFFFF", highlightthickness=2, highlightbackground=DARK_BLUE     ).pack(pady=24)


    # Form Section
    form_frame = tk.Frame(frame,bg=WHITE, highlightthickness=1, highlightbackground=DARK_BLUE)
    form_frame.pack(pady=5, padx=10)

    # Ensure workbook and sheets exist
    if not os.path.exists(EXCEL_FILE):
        wb = openpyxl.Workbook()
        wb.create_sheet(STUDENT_SHEET)
        wb.create_sheet(CATEGORY_SHEET)
        wb.create_sheet(RECORD_SHEET)
        if "Sheet" in wb.sheetnames:
            del wb["Sheet"]
        wb.save(EXCEL_FILE)

    # Load student list from Excel
    def load_student_list():
        student_list = []
        wb = openpyxl.load_workbook(EXCEL_FILE)
        if STUDENT_SHEET in wb.sheetnames:
            sheet = wb[STUDENT_SHEET]
            for row in sheet.iter_rows(min_row=2, values_only=True):
                if row and row[1]:  # Expecting student ID in column 0 and name in column 1
                    student_list.append({"id": row[0], "name": row[1]})
        return student_list

    # Load category list from Excel
    def load_category_list():
        category_list = []
        wb = openpyxl.load_workbook(EXCEL_FILE)
        if CATEGORY_SHEET in wb.sheetnames:
            sheet = wb[CATEGORY_SHEET]
            for row in sheet.iter_rows(min_row=2, values_only=True):
                if row and row[0]:
                    category_list.append(row[0])
        return category_list

    # Initialize comboboxes first
    style = ttk.Style()
    style.theme_use("default")
    style.configure(
    "Custom.TCombobox",
    fieldbackground=WHITE,
    background=WHITE,
    foreground=DARK_BLUE,
    bordercolor=BLUE,
    lightcolor=BLUE,
    darkcolor=BLUE,
    borderwidth=2,
    relief="flat",
    font=("Segoe UI", 13)
)

    combo_student = ttk.Combobox(
    form_frame,
    state="readonly",
    width=28,
    font=("Segoe UI", 13),
    style="Custom.TCombobox"
)
    combo_student.grid(row=0, column=1, pady=6, padx=10)

    combo_category = ttk.Combobox(
    form_frame,
    state="readonly",
    width=28,
    font=("Segoe UI", 13),
    style="Custom.TCombobox"
)
    combo_category.grid(row=2, column=1, pady=6, padx=10)

    label_opts = {
    "font": ("Segoe UI", 13, "bold"),
    "fg": DARK_BLUE,
    "bg": WHITE,
}

    tk.Label(form_frame, text="Student Name:", **label_opts).grid(row=0, column=0, sticky="e", padx=10, pady=6)
    tk.Label(form_frame, text="Amount Paid:", **label_opts).grid(row=1, column=0, sticky="e", padx=10, pady=6)
    tk.Label(form_frame, text="For:", **label_opts).grid(row=2, column=0, sticky="e", padx=10, pady=6)

    entry_amount = tk.Entry( form_frame,width=27,font=("Segoe UI", 14),fg="#17355A", bg=BG_ENTRY,highlightthickness=2,highlightbackground="#4190F3",relief=tk.FLAT,bd=0)
    entry_amount.grid(row=1, column=1, pady=6, padx=8)


    # Function to update student combobox with the latest student list
    def update_student_combobox():
        student_list = load_student_list()
        student_names = [f"{student['name']} (ID: {student['id']})" for student in student_list]
        combo_student['values'] = student_names
        return student_list

    # Function to update category combobox with the latest category list
    def update_category_combobox():
        category_list = load_category_list()
        combo_category['values'] = category_list
        return category_list

    # Load student and category data for initial display
    student_list = update_student_combobox()
    category_list = update_category_combobox()

    # Table Section amount paid
    table_frame = tk.Frame(frame, bg=WHITE, highlightbackground=DARK_BLUE, highlightthickness=2)
    table_frame.pack(expand=True, fill="both", padx=20, pady=10)

    style = ttk.Style()
    style.theme_use("default")
    style.configure("Treeview.Heading", font=("Segoe UI", 13, "bold"), background=BLUE, foreground=WHITE)
    style.configure("Treeview", font=("Segoe UI", 12), rowheight=28, background=WHITE, fieldbackground=WHITE)

    columns = ("Student ID", "Student Name", "Amount Paid", "Payment Category", "Date & Time")
    tree = ttk.Treeview(table_frame, columns=columns, show="headings", style="Treeview")
    for col in columns:
        tree.heading(col, text=col)
        tree.column(col, width=150, anchor="center")
#Page
    vsb = ttk.Scrollbar(table_frame, orient="vertical", command=tree.yview)
    hsb = ttk.Scrollbar(table_frame, orient="horizontal", command=tree.xview)
    tree.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)
    vsb.pack(side="right", fill="y", padx=(0, 2), pady=2)
    hsb.pack(side="bottom", fill="x", padx=2, pady=(0, 2))
    tree.pack(expand=True, fill="both", padx=2, pady=2)

    # Save record to Excel and update table
    def add_record():
        student_name = combo_student.get().strip()
        amount = entry_amount.get().strip()
        category = combo_category.get().strip()

        if not student_name or not amount or not category:
            messagebox.showerror("Input Error", "All fields are required.")
            return

        try:
            float(amount)
        except ValueError:
            messagebox.showerror("Input Error", "Amount must be a number.")
            return

        now = datetime.now()
        date_time = f"{now.month}/{now.day}/{now.year % 100}, {now.strftime('%I:%M %p')}"

        # Extract student ID
        student_id = student_name.split(" (ID: ")[-1].split(")")[0]
        student_name_display = student_name.split(" (ID: ")[0]

        # Append to Treeview
        tree.insert("", "end", values=(student_id, student_name_display, f"₱{amount}", category, date_time))

        # Append to Excel
        wb = openpyxl.load_workbook(EXCEL_FILE)
        if RECORD_SHEET not in wb.sheetnames:
            wb.create_sheet(RECORD_SHEET)
        sheet = wb[RECORD_SHEET]

        if sheet.max_row == 1:
            sheet.append(["Student ID", "Student Name", "Amount Paid", "Payment Category", "Date & Time"])

        sheet.append([student_id, student_name_display, amount, category, date_time])
        wb.save(EXCEL_FILE)

        # Reset fields
        combo_student.set("")
        entry_amount.delete(0, tk.END)
        combo_category.set("")

    def on_add_enter(e): btn_create.config(bg="#007A17")
    def on_add_leave(e): btn_create.config(bg=GREEN)

    btn_create = tk.Button(
        form_frame,
        text="Create Permanent Record",
        font=("Segoe UI", 13, "bold"),
        bg=GREEN, fg=WHITE,
        activebackground="#007A17", activeforeground=WHITE,
        bd=0, relief=tk.FLAT, cursor="hand2",
        command=add_record
    )
    btn_create.grid(row=3, column=0, sticky="nsew", columnspan=2, padx=5, pady=8)
    btn_create.bind("<Enter>", on_add_enter)
    btn_create.bind("<Leave>", on_add_leave)

    # Load existing payment records from Excel
    def load_existing_records():
        tree.delete(*tree.get_children())  # Clear existing entries
        wb = openpyxl.load_workbook(EXCEL_FILE)
        if RECORD_SHEET in wb.sheetnames:
            sheet = wb[RECORD_SHEET]
            for row in sheet.iter_rows(min_row=2, values_only=True):
                if row and len(row) >= 5:
                    tree.insert("", "end", values=(row[0], row[1], f"₱{row[2]}", row[3], row[4]))

    load_existing_records()

    # Function to refresh data every 0.5 second
    def auto_refresh():
        load_existing_records()           # Reload the records
        update_student_combobox()         # Reload student list
        update_category_combobox()        # Optional: reload category too
        frame.after(500, auto_refresh)    # Refresh every 500 milliseconds (0.5 second)

    auto_refresh()  # Start auto-refresh

    return frame
