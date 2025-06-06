import tkinter as tk
from tkinter import ttk, messagebox
from openpyxl import load_workbook, Workbook
import os
from openpyxl import Workbook
from bg_images import *
from bg_loader import load_bg_image

BLUE = "#4190F3"
YELLOW = "#F5B940"
DARK_BLUE = "#17355A"
GREEN = "#009820"
WHITE = "#FFFFFF"
BG_ENTRY = "#F7FAFC"

student_list = []
record_list = []
required_payments = {}
last_modified_time = None  # Track the last modification time of the file

def ensure_excel_file_structure(excel_path):
    if not os.path.exists(excel_path):
        # Create a new workbook with the required sheets
        wb = Workbook()
        wb.remove(wb.active)  # Remove the default sheet

        wb.create_sheet("Student_Management")
        wb.create_sheet("Payment_Records")
        wb.create_sheet("Payment_Categories")

        wb.save(excel_path)
    else:
        wb = load_workbook(excel_path)
        modified = False

        required_sheets = ["Student_Management", "Payment_Records", "Payment_Categories"]
        for sheet_name in required_sheets:
            if sheet_name not in wb.sheetnames:
                wb.create_sheet(sheet_name)
                modified = True

        if modified:
            wb.save(excel_path)


def load_student_status_data(excel_path):
    global student_list, record_list, required_payments
    student_list.clear()
    record_list.clear()
    required_payments.clear()

    try:
        wb = load_workbook(excel_path)
    except Exception as e:
        messagebox.showerror("Error", f"Failed to open Excel file: {e}")
        return

    if 'Student_Management' not in wb.sheetnames or 'Payment_Records' not in wb.sheetnames or 'Payment_Categories' not in wb.sheetnames:
        messagebox.showerror("Error", "Required sheets not found in Excel (Student_Management, Payment_Records, Payment_Categories).")
        return

    # Load students
    student_sheet = wb['Student_Management']
    for row in student_sheet.iter_rows(min_row=2, values_only=True):
        if row[0] is None: continue
        student_id, first_name, middle_initial, last_name = row
        full_name = f"{first_name} {last_name}"
        student_list.append({"id": student_id, "name": full_name})

    # Load required payments (payment categories)
    category_sheet = wb['Payment_Categories']
    for row in category_sheet.iter_rows(min_row=2, values_only=True):
        if row[0] is None: continue
        category, required_amount = row
        required_payments[category.lower()] = float(required_amount or 0)

    # Load payment records
    payment_sheet = wb['Payment_Records']
    for row in payment_sheet.iter_rows(min_row=2, values_only=True):
        if len(row) != 5:
            print(f"Skipping invalid row in Payment_Records: {row}")
            continue
        student_id, student_name, amount_paid, category, _ = row
        if not student_id or not category:
            continue
        record_list.append({
            "student_id": student_id,
            "student_name": student_name,
            "amount_paid": float(amount_paid or 0),
            "payment_category": category.lower()
        })

    global last_modified_time
    last_modified_time = os.path.getmtime(excel_path)  # Store last modified time

def create_student_status_page(parent):
    frame = tk.Frame(parent,bg=WHITE, highlightthickness=2, highlightbackground=DARK_BLUE)
    frame.pack(fill="both", expand=True)
    
    canvas = tk.Canvas(frame,bg=WHITE, highlightthickness=2, highlightbackground=DARK_BLUE)
    canvas.place(x=0, y=0, relwidth=1, relheight=1)  # Cover the entire frame area
    load_bg_image(canvas, STUDENT_STATUS_BACKGROUND_IMAGE)  # Load the background image

    tk.Label(frame, text="Student Status Page", font=("Segoe UI", 32, "bold"),fg="#4190F3",  bg="#FFFFFF", highlightthickness=2, highlightbackground=DARK_BLUE).pack(pady=24)

    # Search bar
    def search_student(event=None):
        query = search_entry.get().strip().lower()
        tree.delete(*tree.get_children())  # Clear the current treeview content

        if not query:  # If search entry is empty, display all students
            for student in student_list:
                insert_student_row(student)
        else:
            match_found = False
            for student in student_list:
                if query in student["name"].lower():
                    insert_student_row(student)
                    match_found = True

    search_frame = tk.Frame(frame,bg=WHITE, highlightthickness=2, highlightbackground=DARK_BLUE)
    search_frame.pack(pady=5, fill="x", padx=20)

    tk.Label(
        search_frame,
        text="Search Student Name:",
        font=("Arial", 12, "bold"),
        fg="#17355A"
    ).pack(side="left", padx=8, pady=6)

    search_entry = tk.Entry(
        search_frame,
        width=50,
        font=("Arial", 12),
        fg="#17355A",         # Dark blue text
        bg="#F7FAFC",         # Light background
        highlightthickness=2,
        highlightbackground="#4190F3",  # Blue border
        relief=tk.FLAT,
        bd=0
    )
    search_entry.pack(side="left", fill="x", expand=True, padx=8, pady=6)

    # Bind the key release event to trigger search
    search_entry.bind("<KeyRelease>", search_student)

    # Table
    table_frame = tk.Frame(frame)
    table_frame.pack(expand=True, fill="both", padx=20, pady=10)

    # Function to create columns dynamically based on payment categories
    def update_columns():
        columns = ["Student ID", "Student Name"] + [cat.title() for cat in required_payments.keys()]
        tree.configure(columns=columns)

        for col in columns:
            tree.heading(col, text=col)
            tree.column(col, anchor="center", stretch=True)

    tree = ttk.Treeview(table_frame, show="headings")

    # Create columns initially
    update_columns()

    vsb = ttk.Scrollbar(table_frame, orient="vertical", command=tree.yview)
    hsb = ttk.Scrollbar(table_frame, orient="horizontal", command=tree.xview)
    tree.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)

    vsb.pack(side="right", fill="y")
    hsb.pack(side="bottom", fill="x")
    tree.pack(fill="both", expand=True)

    def insert_student_row(student):
        student_records = [r for r in record_list if r["student_id"] == student["id"]]
        payments_by_category = {cat: 0.0 for cat in required_payments.keys()}
        for record in student_records:
            cat = record["payment_category"]
            if cat in payments_by_category:
                payments_by_category[cat] += record["amount_paid"]

        row_data = [student["id"], student["name"]]
        for cat in required_payments.keys():
            paid = payments_by_category[cat]
            required = required_payments[cat]
            cell_value = f"{paid:.0f}/{required:.0f}" if required > 0 else ""
            row_data.append(cell_value)

        tree.insert("", "end", values=row_data)

    # Load all students initially
    for student in student_list:
        insert_student_row(student)

    def check_for_updates():
        global last_modified_time
        current_modified_time = os.path.getmtime("userdata.xlsx")
        if current_modified_time != last_modified_time:
            # Reload the data and update columns
            load_student_status_data("userdata.xlsx")

            # Update columns and table data
            update_columns()

            tree.delete(*tree.get_children())  # Clear existing rows
            for student in student_list:
                insert_student_row(student)  # Re-insert updated rows

            last_modified_time = current_modified_time  # Update the last modified time

        # Check for updates every 1000 ms (1 second)
        frame.after(10, check_for_updates)

    # Start the update check loop
    check_for_updates()

    return frame

# Initial load of data
ensure_excel_file_structure("userdata.xlsx")
load_student_status_data("userdata.xlsx")